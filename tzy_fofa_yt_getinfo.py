# author:soufaker
# time:2022/06/10

import requests
import time
import optparse
import json
import base64
import httpx
import urllib.request
import re
import urllib3
import tldextract
import threading
from urllib.parse import quote
from lxml import etree
from openpyxl import Workbook
from configparser import ConfigParser

def beianchaxun(url):
    urls = url
    URL = ("https://beian.tianyancha.com/search/{}".format(urls))
    headers = {
        'Connection': 'close',
        'Cache-Control': 'max-age=0',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'zh-CN,zh;q=0.8',
    }
    try:
        req = requests.get(url=URL, headers=headers)
        html = req.content
        titlere = r'<span class="ranking-ym" rel="nofollow">(.+?)</span>'
        title = re.findall(titlere, html.decode('utf-8'))
        if len(title) == 0:
            print(urls, "该公司天眼查找不到备案域名", '\n')
        else:
            print('\n', urls, "备案域名: ")
            for i in title:
                if i not in black_domian and i != '':
                    all_domain_list.append(i)
                    print(i)
    except:
        print('有异常，无法爬取')

# 选取想要的域名元素
def Get_MiddleStr(content, startStr, endStr, num):  # 获取中间字符串的⼀个通⽤函数
    try:
        startIndex = content.index(startStr)
        if startIndex >= 0:
            if num != 1:
                endIndex = startIndex
                startIndex = endIndex + num

            else:
                startIndex += len(startStr)
                content2 = content[startIndex:]
                endIndex = content2.index(endStr)+startIndex

        return content[startIndex:endIndex]

    except:
        return 'ok'


# 获取公司查询ID
def Get_Company_Id(company_name, company_name_list):
    company_id_list = []
    url = "https://www.tianyancha.com/search?key="
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    if company_name != None:
        try:
            url = url + quote(company_name, 'utf-8')
            print(url)
            response = requests.get(url=url, headers=header).content
            html = etree.HTML(response)
            print()
            path = '/html/body/div[1]/div[2]/div/div[2]/section/main/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[1]/div[1]/a//@href'
            company_id = html.xpath(path)[0][35:]
            print(company_id)
            company_id_list.append(company_id)
        except Exception as e:
            print(e)
            print(f'error file:{e.__traceback__.tb_frame.f_globals["__file__"]}')
            print(f"error line:{e.__traceback__.tb_lineno}")

    else:
        with open(company_name_list, 'r', encoding='utf-8') as f:
            for company in f:
                try:
                    response = requests.get(url=url + quote(company, 'utf-8'), headers=header).content
                    html = etree.HTML(response)
                    path = '//a/@href'
                    company_id = html.xpath(path)[4][35:]
                    company_id_list.append(company_id)
                except:
                    print('访问过于频繁,请切换IP或稍后再试!')


    return company_id_list




# 获取公司占百分百股权且未注销的子公司域名信息
def Get_ALL_Sub_Cmpany_Domain(company_id):
    company_info_list = []
    company_url_list = []
    for main_company in company_id:
        l, l2, l3 = Get_Sub_Company_Domain(main_company)
        if len(l) != 0:
            for i in l:
                company_info_list.append(i)
                print(company_info_list)
        if len(l2) != 0:
            for i2 in l2:
                company_url_list.append(i2)
                print(company_url_list)

        com_id = []
        com_id.append(main_company)
        while com_id:
            co_id = com_id.pop(0)
            company_id_list = get_all_page_id(co_id)
            if len(company_id_list) != 0:
                for co in company_id_list:
                    com_id.append(co)
                for ci in company_id_list:
                    l, l2, l3 = Get_Sub_Company_Domain(ci)
                    if len(l) != 0:
                        for i in l:
                            company_info_list.append(i)
                    if len(l2) != 0:
                        for i2 in l2:
                            company_url_list.append(i2)
                    if len(l3) != 0:
                        for l in l3:
                            com_id.append(l)

    return company_info_list, company_url_list


def Get_Sub_Company_Domain(company_id):
    company_info_list = []
    company_url_list = []
    err_list = []
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    with open('./config/cookie.ini', 'r', encoding='utf-8') as co:
        f = co.read()
        f2 = f.replace(' ', '')
        header['Cookie'] = f2[6:]
    print(company_id)
    company_url = "https://www.tianyancha.com/company/" + str(company_id)
    print(company_url)
    try:
        res = requests.get(url=company_url, headers=header).text
        resp = res.replace(' ', '').replace('\n', '').replace('\t', '').replace("\"", "")
        try:
            name = Get_MiddleStr(resp, '<title>', "-天眼查", 1)
            url = Get_MiddleStr(resp, "_blankrel=nofollownoreferrer>", "</a>", 1)
            email = Get_MiddleStr(resp, "<spanclass=index_detail-email__B_1Tq>", "</span><", 1)
            mobile = Get_MiddleStr(resp,'电话：</span><span><spanclass=index_detail-tel__fgpsE>',"</span>",1)

            if '暂无网址' in res:
                url = '暂无网址'
            elif len(email) < 1 or email == 'ok':
                email = '暂无邮箱'
            elif len(mobile) < 1 or email == 'ok':
                mobile = '暂无电话'

            info = [name, url, email, mobile]
            company_info_list.append(info)
            if url != '暂无网址':
                company_url_list.append(url)
        except Exception as e:
            print(e)
            print(f'error file:{e.__traceback__.tb_frame.f_globals["__file__"]}')
            print(f"error line:{e.__traceback__.tb_lineno}")

    except Exception as e:
        print(e)
        print(f'error file:{e.__traceback__.tb_frame.f_globals["__file__"]}')
        print(f"error line:{e.__traceback__.tb_lineno}")

    return company_info_list, company_url_list, err_list

def get_all_page_id(id):
    id_list = []
    header = {
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:104.0) Gecko/20100101 Firefox/104.0",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
    }
    with open('./config/cookie.ini', 'r', encoding='utf-8') as co:
        f = co.read()
        f2 = f.replace(' ', '')
        header['Cookie'] = f2[6:]

    company_url = "https://capi.tianyancha.com/cloud-company-background/company/investListV2?_=1663738376979"
    data = '{"gid":"'+str(id)+'","pageSize":100,"pageNum":1,"province":"","percentLevel":"-100","category":"-100"}'
    header = {
        "Content-Type": "application/json",
        "Cookie": "PHPSESSID=soh8vb6igolqj31r4kpp6nsor4; SERVERID=ae57de1396dcdc6bb63529ba5709f96f|1660720639|1660713694; acw_tc=276077e116607206371748293e7ea3827bf3732b1f206600ba66a7a66c4c09",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:104.0) Gecko/20100101 Firefox/104.0",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
    }

    response = requests.post(url=company_url, headers=header, data=data).text
    j = json.loads(response)
    print(j)
    for i in range(len(j['data']['result'])):
        l1 = []
        occ_r = j['data']['result'][i]['percent'][:-1]
        print(occ_r)
        if occ_r != '':
            occ_rev = float(occ_r)
            if occ_rev >= float(company_occ):
                l1.append(j['data']['result'][i]['id'])
                # l1.append(j['data']['result'][i]['regStatus'])
                id_list.append(l1[0])

    return id_list


def Write_To_Excel(company_info_list,all_info_list):
    t = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
    wb = Workbook()
    ws1 = wb.create_sheet('天眼查基本信息', 0)
    ws2 = wb.create_sheet('收集域名信息', 0)

    ws1['A1'] = '公司名'
    ws1['B1'] = '网址'
    ws1['C1'] = '邮箱'
    ws1['D1'] = '联系电话'
    for l in company_info_list:
        ws1.append(list(l))

    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 36
    ws1.column_dimensions['C'].width = 36
    ws1.column_dimensions['D'].width = 36

    ws2['A1'] = '网址'
    ws2['B1'] = '状态码'
    ws2['C1'] = '网页标题'
    for l in all_info_list:
        ws2.append(l)

    ws2.column_dimensions['A'].width = 36
    ws2.column_dimensions['B'].width = 36
    ws2.column_dimensions['C'].width = 36

    wb.save("暴露面收集" + t + ".xlsx")


def Write_To_Txt(company_url_list, txt):
    t = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
    with open(txt + 'all_url_list' +'_'+ t + '.txt', 'w+', encoding='utf-8') as f:
        for c in list(set(company_url_list)):
            f.writelines(str(c) + '\n')


def Get_icp_Num(company_name):
    company_url = "https://beian.tianyancha.com/search/" + company_name
    print(company_url)
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    try:
        response = requests.get(url=company_url, headers=header).text
        resp = response.replace(' ', '').replace('\n', '').replace('\t', '').replace("\"", "")

        flag = True
        page_num = 1

        while flag:
            for i in range(2, 100):
                str_flag = '/p' + str(i)
                if str_flag in resp:
                    flag = True
                    page_num += 1
                else:
                    flag = False
    except:
        page_num = 0

    return page_num


def get_icp_name_list(company_name, page_num):
    all_icp_list = []
    for i in range(1, page_num + 1):
        header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        if i == 1:
            company_url = "https://beian.tianyancha.com/search/" + str(company_name)
        else:
            company_url = "https://beian.tianyancha.com/search/" + str(company_name) + '/p' + str(i)

        res = requests.get(url=company_url, headers=header)
        response2 = res.content
        html = etree.HTML(response2)
        path = "/html/body/div[2]/div/div[2]/div[1]/div[2]/div[2]/table/tbody/tr/td[2]/a/text()"
        icp_name_list = html.xpath(path)

        for icp in icp_name_list:
            all_icp_list.append(icp)

    return all_icp_list


def yt_get_info(name_list):
    all_url_list = []
    for name in name_list:
        search_key = 'domain=' + name
        keyword = base64.urlsafe_b64encode(search_key.encode("utf-8"))  # 把输入的关键字转换为base64编码
        page = 0
        num = 0
        api_num = 0
        while True:
            try:
                page += 1
                api_num += 1
                url = "https://hunter.qianxin.com/openApi/search?api-key={}&search={}&page={}&page_size=100&is_web=1&status_code=200".format(
                    hunter_config_list[api_num], keyword.decode(), page)
                print(url)
                r = requests.get(url, timeout=3)
                res = json.loads(r.text)
                jf = res['data']['rest_quota'][7:]
                if hunter_config_list[api_num] == '6e74c5a2bde8ae575057befad48c727923c472d5cd4811b23e893666a370f16c':
                    tump_jf = jf
                    jf = int(tump_jf) - 5521

                print('当前api剩余每日免费积分:' + str(jf))
                if int(jf) < 0:
                    if int(api_num) > int(len(hunter_config_list)):
                        api_num += 1
                        url = "https://hunter.qianxin.com/openApi/search?api-key={}&search={}&page={}&page_size=100&is_web=1&status_code=200".format(
                            hunter_config_list[api_num], keyword.decode(), page)
                        print(url)

                    else:
                        print('所有api积分为0,请明日在尝试')

                r = requests.get(url, timeout=3)
                res = json.loads(r.text)
                jf = res['data']['rest_quota'][7:]
                print('当前api剩余每日免费积分:' + jf)
                loadurl = res['data']['arr']
                for i in loadurl:
                    all_url_list.append(i['url'])
                    num += 1
                    print(i['url'])

            except Exception as e:
                print("\n")
                print("采集完成，共采集{}条Url".format(num))
                break
    return all_url_list


def get_fofa_url(domain):
    all_url_list = []
    search_key = "domain=" + domain
    search_data_b64 = base64.b64encode(search_key.encode("utf-8")).decode("utf-8")
    search = 'https://fofa.info/api/v1/search/all?email=' + fofa_email + '&size=1000' + '&key=' + fofa_key + '&qbase64=' + search_data_b64
    print(search)
    while True:
        try:
            r = requests.get(search, timeout=30)
            res = json.loads(r.text)
            size = len(res['results'])
            for i in range(0, int(size)):
                result = res['results'][i][0]
                print(result)
                all_url_list.append(result)
            break

        except:
            print('fofa连接超时,正在重试！')
            time.sleep(60)
            pass
        continue

    return list(set(all_url_list))


# 检测存活
class myThreade(threading.Thread):
    def __init__(self, all_url_list):
        threading.Thread.__init__(self)
        self.all_url_list = all_url_list

    def get_code_method(slef, url):
        header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3298.4 Safari/537.36'
        }

        try:
            res = httpx.get(
                url, headers=header, verify=False, timeout=1
            )
            code = res.status_code
            print('{} {}'.format(code, url))
            return code
        except:
            print(url + '该网站无法访问!')

    # 获取网站标题
    def get_title(self,url):
            try:
                print(url)
                page = urllib.request.urlopen(url=url,timeout=60)
                html = page.read().decode('utf-8')
                title = re.findall('<title>(.+)</title>', html)
                print(title)
                return title[0]
            except:
                return ''

    def get_code_status_list(self, urls_list):
        """
        检测urls状态码
        :param urls_file: urls文件
        :return: 状态码
        verify: 取消证书认证
        """
        url_list = []

        for url in urls_list:
            url_l = []
            disable_warnings()
            code = self.get_code_method(url)
            title = self.get_title(url)
            url_l.append(url)
            url_l.append(code)
            url_l.append(title)
            url_list.append(url_l)

        return url_list

    def run(self):
        all_url_list1 = []
        for url in self.all_url_list:
            if 'http' not in url:
                for i in range(0, 2):
                    if i == '0':
                        all_url_list1.append('http://' + url)
                    else:
                        all_url_list1.append('https://' + url)
            else:
                all_url_list1.append(url)

        url_lists = self.get_code_status_list(set(all_url_list1))

        for url in url_lists:
            all_info_list.append(url)


def disable_warnings():
    """
    解除去掉证书后总是抛出异常告警
    """
    urllib3.disable_warnings()


def split_list_average_n(origin_list, n):
    for i in range(0, len(origin_list), n):
        yield origin_list[i:i + n]


def get_all_url_fo_yt(company_info_list,company_domains):
    # 查询所有域名
    for com in company_info_list:
        beianchaxun(com[0])
    # 查询
    for com in company_info_list:
        if '.' in com[1]:
            val = tldextract.extract(com[1])
            if val.registered_domain != '':
                all_domain_list.append(val.registered_domain)

    # 调用域名进行查询
    domains_list = company_domains.split(",")

    for domain in domains_list:
        all_domain_list.append(domain)

    new_domain_list = []
    for domain in all_domain_list:
        for ch in domain:
            if u'\u4e00' <= ch <= u'\u9fff':
                print('存在中文')
                continue
        domain1 = domain.replace('<em>ok</em>', '')
        if domain1 not in black_domian and domain != '' and '<' not in domain:
            new_domain_list.append(domain1)

    print('当前搜集了如下域名')
    print(new_domain_list)

    # 调用鹰图,并添加到所有搜集的列表
    print('开始调用鹰图')
    for domain in new_domain_list:
        url_list = yt_get_info(domain)
        for url in url_list:
            all_url_list.append(url)

    print('开始使用fofa查询')
    for domain in list(set(new_domain_list)):
        url_list = get_fofa_url(str(domain))

        for url in url_list:
            all_url_list.append(url)
    last_all_url_list = []

    for url in all_url_list:
        if url not in last_all_url_list:
            last_all_url_list.append(url)

    return last_all_url_list



def jxch(all_url_list):
    threads = []
    # 线程数
    print('开始检测存活')
    thread_count = 20

    count = len(all_url_list) // thread_count
    if count == 0:
        count = 1

    co = split_list_average_n(all_url_list, count)

    for i in co:
        threads.append(myThreade(i))
    for t in threads:
        t.start()
    for t in threads:
        t.join()

if __name__ == '__main__':
    # 参数设置
    parser = optparse.OptionParser()
    parser.add_option('-c', '--company', action='store', dest="company_name")
    parser.add_option('-l', '--list', action='store', dest="company_name_list")
    parser.add_option('-o', '--occ', action='store', default='100', dest="company_occ")
    parser.add_option('-d', '--domain', action='store', default='', dest="company_domains")
    options, args = parser.parse_args()

    # 加载配置文件
    cf = ConfigParser()
    cf.read('./config/config.ini')
    global hunter_config_list
    hunter_config_list = []
    global fofa_key
    fofa_key = ''
    global fofa_email
    fofa_email = ''
    global black_domian
    black_domian = []

    c_len = cf.options('hunter')
    for i in c_len:
        hunter_config_list.append(cf.get('hunter', i))
    fofa_key = cf.get('fofa', 'fofa_key')
    fofa_email = cf.get('fofa','fofa_email')
    black = cf.get('black_domain','domain')
    black_domian = black.split(',')
    # 所有搜集到的URL列表
    all_url_list = []

    # 最后整理出的url列表
    all_url_list2 = []

    # 全局收集的信息
    global all_info_list
    all_info_list = []

    # 全局搜寻的域名
    all_domain_list = []

    threadLock = threading.Lock()

    # 参数获取
    global company_occ
    company_name = options.company_name
    company_domains = options.company_domains
    company_list = options.company_name_list
    company_occ = float(options.company_occ)
    company_id = Get_Company_Id(company_name, company_list)
    company_info_list, company_url_list = Get_ALL_Sub_Cmpany_Domain(company_id)


    all_url_list = get_all_url_fo_yt(company_info_list,company_domains)

    print(all_url_list)
    jxch(all_url_list)

    Write_To_Txt(list(set(all_domain_list)), 'all_domain')
    Write_To_Excel(list(set(company_info_list)),list(set(all_info_list)))