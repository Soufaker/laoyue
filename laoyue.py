# encoding:utf-8
# author:Soufaker
# time:2023/12/19
import argparse
import random
import traceback
import requests
import time
import optparse
import json
import base64
import urllib.request
import re
import tldextract
from dingtalkchatbot.chatbot import DingtalkChatbot
from datetime import datetime
import math
from openpyxl import Workbook
from configparser import ConfigParser
import subprocess

requests.packages.urllib3.disable_warnings()
import urllib
import os


def Write_To_Excel(all_info_list, mgwj_list, ld_list, httpx_info, fs_list, host_scan_list):
    t = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
    wb = Workbook()
    ws1 = wb.create_sheet('收集域名信息', 0)
    ws2 = wb.create_sheet('fscan扫描信息', 0)
    ws3 = wb.create_sheet('敏感信息', 0)
    ws4 = wb.create_sheet('漏洞信息', 0)
    ws5 = wb.create_sheet('host碰撞信息', 0)
    ws6 = wb.create_sheet('httpx信息', 0)

    ws1['A1'] = '网址'
    ws1['B1'] = 'ip地址'
    ws1['C1'] = '端口'
    ws1['D1'] = '网页标题'
    ws1['E1'] = '协议'
    ws1['F1'] = '状态码'
    ws1['G1'] = '框架'
    ws1['H1'] = '备案号'
    for l in all_info_list:
        ws1.append(l)

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 7
    ws1.column_dimensions['D'].width = 25
    ws1.column_dimensions['E'].width = 7
    ws1.column_dimensions['F'].width = 7
    ws1.column_dimensions['G'].width = 30
    ws1.column_dimensions['H'].width = 25

    ws2['A1'] = '漏洞地址'
    ws2.column_dimensions['A'].width = 70
    for l in fs_list:
        list1 = []
        list1.append(l)
        ws2.append(list1)

    ws3['A1'] = '地址'
    ws3['B1'] = '状态码'
    ws3['C1'] = '大小'

    ws3.column_dimensions['A'].width = 70
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 15

    for l in mgwj_list:
        ws3.append(l)

    ws4['A1'] = '漏洞名字'
    ws4['B1'] = '漏洞等級'
    ws4['C1'] = '漏洞地址'

    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 10
    ws4.column_dimensions['C'].width = 50

    for l in ld_list:
        ws4.append(l)

    ws5['A1'] = '碰撞成功信息'

    ws5.column_dimensions['A'].width = 150

    for l in host_scan_list:
        a = []
        a.append(l)
        ws5.append(a)

    ws6['A1'] = '网址'
    ws6['B1'] = '状态码'
    ws6['C1'] = '标题'

    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 10
    ws6.column_dimensions['C'].width = 50

    for l in httpx_info:
        ws6.append(l)

    wb.save("./result/baolumian/暴露面收集" + t + ".xlsx")


def yt_info(url):
    temp_url_list = []
    temp_url_list.append(url)
    while temp_url_list:
        temp_url_list.pop()
        r = requests.get(url, timeout=3)
        res = json.loads(r.text)
        if str(res['code']) == '429':
            temp_url_list.append(url)
            time.sleep(1)
            continue
        loadurl = res['data']['arr']

        if loadurl != None:
            company = loadurl[0]['company']
            print(company)
            for i in range(0, len(loadurl)):
                domain = loadurl[i]['domain']
                if '.' in domain:
                    val = tldextract.extract(domain)
                    if val.registered_domain not in all_domain_list:
                        continue
                info = []
                arr_url = loadurl[i]['url']
                # if '中国' in loadurl[i]['isp']:
                #     arr_ip = loadurl[i]['ip']
                # else:
                #     arr_ip = '存在CDN:' + loadurl[i]['ip']
                arr_ip = isCDN(loadurl[i]['domain'], loadurl[i]['ip'])
                arr_port = loadurl[i]['port']
                arr_web_title = loadurl[i]['web_title']
                arr_protocol = loadurl[i]['protocol'] + ',' + loadurl[i]['base_protocol']
                arr_status_code = loadurl[i]['status_code']
                arr_component = loadurl[i]['component']
                arr_beianhhao = loadurl[i]['number']
                if arr_beianhhao == '':
                    arr_beianhhao = '-'
                arr_all_component = ''
                if arr_component != None:
                    for i in arr_component:
                        s = i['name'] + i['version']
                        arr_all_component = arr_all_component + '|' + s
                info.append(arr_url)
                info.append(arr_ip)
                info.append(arr_port)
                info.append(arr_web_title)
                info.append(arr_protocol)
                info.append(arr_status_code)
                info.append(arr_all_component)
                info.append(arr_beianhhao)
                # print(info)
                flag = True
                for b in black_domian:
                    if b in arr_url:
                        flag = False
                if flag:
                    all_info_list.append(info)
            return 1
        else:
            return 2


def yt_get_info(name_list):
    new_list = fy_list(name_list, hunter_count)
    for domain_list in new_list:
        try:
            domain_all = ''
            for domain in domain_list:
                if domain != '':
                    if isIP(domain):
                        domain_all = domain_all + "ip=" + domain + '||'
                    else:
                        domain_all = domain_all + "domain=" + domain + '||'
            print(domain_all)
            search_key = '(' + domain_all[0:-2] + ')' + str(fofa_keyword)
            keyword = base64.urlsafe_b64encode(search_key.encode("utf-8"))  # 把输入的关键字转换为base64编码
            page = 1
            api_num = 0
            while True:
                # 测试第一个API积分是否够用
                url = "https://hunter.qianxin.com/openApi/search?api-key={}&search={}&page={}&page_size=1&is_web=1".format(
                    hunter_config_list[api_num], keyword.decode(), page)
                r = requests.get(url)
                res = json.loads(r.text)

                if str(res['code']) == '429':
                    continue
                if str(res['code']) == '401':
                    if int(api_num) < int(len(hunter_config_list)):
                        api_num += 1
                    continue
                if str(res['code']) == '40204' or str(res['code']) == '40201':
                    if int(api_num) < int(len(hunter_config_list)):
                        api_num += 1
                        print('上一个积分已经用完,切换第' + str(int(api_num) + 1) + '个API')
                        continue

                url = "https://hunter.qianxin.com/openApi/search?api-key={}&search={}&page={}&page_size={}&is_web=1".format(
                    hunter_config_list[api_num], keyword.decode(), page, yt_size)
                print(url)
                pd_num = yt_info(url)
                if pd_num == 2:
                    print('未查到数据')
                break
        except Exception as e:
            traceback.print_exc()
            print('积分清0退出循环', e)
            break


def get_title(url):
    try:
        page = urllib.request.urlopen(url=url, timeout=0.5)
        html = page.read().decode('utf-8')
        title = re.findall('<title>(.+)</title>', html)
        return title[0]
    except:
        return ''


def isCDN(domain, ip):  # 判断目标是否存在CDN
    parm = 'nslookup ' + domain
    try:
        result = os.popen(parm).read()
        l = result.split('Name:')
        if result.count("Name") > 1 or domain not in l[1] or 'gslb' in result or 'dns' in result or 'cache' in result:
            return "存在CDN" + str(ip)
        else:
            if ip not in ip_list:
                ip_list.append(ip)
            # 添加到对应域名IP文件列表里,方便进行host碰撞
            all_domain_ip_list.append(domain + '-' + ip)
            return ip
    except:
        return '-'


def fy_list(list1, count):
    new_list = []
    num = len(list1)
    if int(num) > int(count):
        n = num // int(count)
    else:
        n = 1

    for i in range(0, n):
        one_list = list1[math.floor(i / n * num):math.floor((i + 1) / n * num)]
        new_list.append(one_list)
    return new_list


def get_fofa_url(domain_l):
    new_list = fy_list(domain_l, fofa_count)
    while True:
        for domain_list in new_list:
            try:
                domain_all = ''
                for domain in domain_list:
                    if domain != '':
                        if isIP(domain):
                            domain_all = domain_all + "ip=" + domain + '||'
                        else:
                            domain_all = domain_all + "domain=" + domain + '||'
                print(domain_all)
                search_key = '(' + domain_all[0:-2] + ')' + str(fofa_keyword)
                search_data_b64 = base64.b64encode(search_key.encode("utf-8")).decode("utf-8")
                search = 'https://fofa.info/api/v1/search/all?email=' + fofa_email + '&size=' + fofa_size + '&key=' + fofa_key + '&qbase64=' + search_data_b64 + "&fields=host,ip,port,title,protocol,header,server,product,icp,domain"
                print(search)
                try:
                    r = requests.get(search, verify=False)
                    res = json.loads(r.text)
                    size = len(res['results'])
                    for i in range(0, int(size)):
                        info = []
                        result = res['results'][i]
                        num = len(result)
                        temp = ''
                        if result[4] == 'unknown':
                            result[4] = 'http'
                        if result[9] not in domain_l:
                            continue
                        for j in range(0, int(num) - 1):
                            if j == 1:
                                ip = isCDN(result[9], result[1])
                                result[1] = ip
                            if j == 0:
                                print(result[0])
                                if 'http' not in result[0][0:5]:
                                    if 'http' not in result[4]:
                                        result[0] = 'https://' + result[0]
                                    else:
                                        result[0] = result[4] + '://' + result[0]
                            if j == 6:
                                temp = result[j]
                                continue
                            if j == 7:
                                result[j] = '|' + result[j] + '|' + temp
                            if j == 8:
                                if result[j] == '':
                                    result[j] = '-'
                            if j == 5:
                                if result[j] == '':
                                    result[j] == '-'
                            if j == 3:
                                result[j] = get_title(result[0])
                                if result[j] == '':
                                    result[j] = '-'

                            if j == 5:
                                result[j] = result[j][9:12]
                            info.append(result[j])
                        print(info)
                        flag = True
                        for b in black_domian:
                            if b in result[0]:
                                flag = False
                        if flag:
                            all_info_list.append(info)
                except:
                    continue
            except:
                print('fofa连接超时,正在重试！')
                time.sleep(1)
                continue
        print('查询结束了')
        break


def split_list_average_n(origin_list, n):
    for i in range(0, len(origin_list), n):
        yield origin_list[i:i + n]


def get_all_url_fo_yt():
    company_domains_list = []
    if '.txt' in company_domain:
        with open(company_domain, 'r', encoding='utf-8') as f:
            l = f.readlines()
            for i in l:
                company_domains_list.append(i.strip('\n'))
    else:
        company_domains_list.append(company_domain)

    if len(company_domains_list) != 0:
        for com in company_domains_list:
            all_domain_list.append(com)

    print('当前搜集了如下域名')
    all_qc_domain_list = list(set(all_domain_list))
    if notauto != True:
        # 调用鹰图,并添加到所有搜集的列表
        if is_hunter == '0':
            print('开始调用鹰图')
            yt_get_info(all_qc_domain_list)
        if is_fofa == '0':
            print('开始使用fofa查询')
            get_fofa_url(all_qc_domain_list)
        if is_subfinder == '0':
            print('开启subfinder子域名扫描')
            run_subfinder(all_qc_domain_list)




# 判断是ip还是域名
def isIP(str):
    p = re.compile('^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')
    if p.match(str):
        return True
    else:
        return False


def save_cache(target_list,subfinder_list):
    yt_fofa_add_list = []
    yt_fofa_add_list2 = []
    sm_add_list = []
    # 读取历史扫描信息
    sm_cache_file_list = open('./caches/sm_cache.txt', 'r', encoding='utf-8').read().split('\n')
    # print('target', target_list)
    # print('sm', sm_cache_file_list)
    # 添加fafa_yt记录的历史域名信息
    if len(target_list) != 0:
        for tar in target_list:
            print(tar[0])
            if tar[0] not in sm_cache_file_list:
                if str(tar[5]) == '200' or str(tar[5]) == '301' or str(tar[5]) == '302' or str(tar[5]) == '201' or str(
                        tar[5]) == '404' or str(tar[5]) == '401' or str(tar[5]) == '405':
                    info = []
                    info.append(str(tar[0]))
                    info.append(str(tar[5]))
                    info.append(str(tar[3]))
                    httpx_info.append(info)

                sm_add_list.append(tar[0])
                str_tar = ''
                for t in tar:
                    str_tar = str_tar + ' | ' + str(t)
                yt_fofa_add_list.append(str_tar)
                yt_fofa_add_list2.append(tar)
    #对比subfinder
    for l in subfinder_list:
        l1 = 'https://' + l
        l2 = 'http://' + l
        if l1 not in sm_cache_file_list and l2 not in sm_cache_file_list:
            sm_add_list.append(l)

    for l in yt_fofa_add_list:
        caches_file = open('./caches/fo_yt_cache.txt', 'a', encoding='utf-8')
        caches_file.write(l + '\n')
        caches_file.close()

    return sm_add_list, yt_fofa_add_list2, sm_cache_file_list


def run_subfinder(all_qc_domain_list):
    for domain in all_qc_domain_list:
        print(f"正在扫描域名: {domain}")
        result = subprocess.run(
            ["./inifile/subfinder/subfinder", "-all", "-d", domain],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True
        )
        subfinder_list.append(result.stdout)



def httpx_naabu_scan(filename, sm_cache_file_list):
    caches_file_list = open('./caches/sm_cache.txt', 'r', encoding='utf-8').read().split('\n')
    caches_file = open('./caches/sm_cache.txt', 'a', encoding='utf-8')
    caches_file2 = './inifile/naabu/cache/cache.txt'

    try:
        filename_temp = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + 'temp_port.txt'
        temp_l = open(filename, 'r', encoding='utf-8').read().split('\n')
        # print(temp_l)
        new_filename_temp = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S",
                                                               time.localtime()) + 'new_temp_port.txt'
        with open(new_filename_temp, 'w') as f:
            for i in temp_l:
                if ':' in i:
                    # print('123')
                    # print(i.split(':')[0])
                    f.writelines(i.split(':')[0] + '\n')
                else:
                    f.writelines(i + '\n')

        filename_filter_name = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S",time.localtime()) + 'all_url_list.txt'
        naabu_sm_file = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S",time.localtime()) + '_all_naabu_list.txt'
        update_cache_and_output_new_lines(caches_file2,new_filename_temp,naabu_sm_file)
        naabu_list = open(naabu_sm_file, 'r', encoding='utf-8').read().split('\n')
        print(naabu_list)
        #如果扫描列表为空就返回1
        if len(naabu_list) < 2:
            return 1

        port_scan = './inifile/naabu/naabu  -l ' + naabu_sm_file + ' -top-ports 1000 -o ' + filename_temp
        # print('2 ' + port_scan)
        os.system(port_scan)  # &> /dev/null
        httpx_filename = filename_temp[0:-4] + '_httpx.txt'
        http_list = open(filename_temp, 'r')
        # print('123'+httpx_filename)
        with open(httpx_filename, 'w+') as f:
            for h in http_list:
                # print(h)
                f.writelines('https://' + h)
                f.writelines('http://' + h)
        http_scan = './inifile/httpx/httpx  -l ' + httpx_filename + ' -mc 200,401,403,404,302,301,500,405,501,502  -title   -status-code  -fr -o  ' + filename_filter_name  # &> /dev/null'
        # print('1 ' + http_scan)
        os.system(http_scan)  # &> /dev/null
        # os.system('rm -rf ' + filename)
        # os.system('rm -rf ' + filename_temp)
        httpx_info_l = open(filename_filter_name, 'r', encoding='utf-8', errors='ignore').read().split('\n')
        httpx_info_list = list(set(httpx_info_l))
        for i in httpx_info_list:
            f = i.split(' ')
            if str(f[0]) not in caches_file_list and str(f[0]) != '':
                info = []
                info.append(str(f[0]))
                if '200' in f[1]:
                    info.append('200')
                else:
                    info.append(str(f[1].split('\x1b')[1].split('m')[1]))
                info.append(str(f[2].split('\x1b')[1][4:]))
                httpx_info.append(info)
                print(info)
        print(httpx_info)

        #
        new_filename_filter_name = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S",
                                                                      time.localtime()) + 'new_all_url_list.txt'
        print('new_filename_filter_name', new_filename_filter_name)

        caches_file_list_1 = open(new_filename_filter_name, 'w+', encoding='utf-8')
        # 写入awvs文件
        file_list = []
        for f in httpx_info:
            if f[0] not in file_list:
                file_list.append(f[0])
        print(file_list)
        # print(sm_cache_file_list)
        for l in file_list:
            if 'http' not in l:
                l1 = 'http://' + l
                l2 = 'https://' + l
                if l2 not in sm_cache_file_list:
                    caches_file.writelines(l2 + '\n')
                    caches_file_list_1.writelines(l2 + '\n')
                if l1 not in sm_cache_file_list:
                    caches_file.writelines(l1 + '\n')
                    caches_file_list_1.writelines(l1 + '\n')
            else:
                if l not in sm_cache_file_list:
                    print('0000000000000000000000')
                    caches_file.writelines(l + '\n')
                    caches_file_list_1.writelines(l + '\n')
        caches_file.close()
        # awvs
        scan_awvs(file_list)

        return new_filename_filter_name
    except Exception as e:
        traceback.print_exc()
        print(e)
        os.system('touch ' + filename_filter_name)
        print('12312312')
        if len(httpx_info) > 0:
            file_list2 = []
            for f in httpx_info:
                file_list2.append(f[0])
            print(file_list2)
            scan_awvs(file_list2)
            caches_file_list_1 = open(filename_filter_name, 'w', encoding='utf-8')
            for l in file_list2:
                if 'http' not in l:
                    l1 = 'http://' + l
                    l2 = 'https://' + l
                    if l2 not in sm_cache_file_list:
                        caches_file.write(l2 + '\n')
                        caches_file_list_1.write(l2 + '\n')
                    if l1 not in sm_cache_file_list:
                        caches_file.write(l1 + '\n')
                        caches_file_list_1.write(l1 + '\n')
                else:
                    if l not in sm_cache_file_list:
                        print('0000000000000000000000')
                        caches_file.write(l + '\n')
                        caches_file_list_1.write(l + '\n')

        return filename_filter_name

#文件对比
def update_cache_and_output_new_lines(file_a, file_b, output_file):
    # 读取文件A的内容到集合
    with open(file_a, 'r') as fa:
        lines_a = set(line.strip() for line in fa)

    # 读取文件B并找出新行
    new_lines = []
    with open(file_b, 'r') as fb:
        for line in fb:
            stripped_line = line.strip()
            if stripped_line not in lines_a:
                new_lines.append(stripped_line)

    # 将新行追加到文件A和输出到新文件
    with open(file_a, 'a') as fa_append, open(output_file, 'w') as out_file:
        for line in new_lines:
            fa_append.write(line + '\n')
            out_file.write(line + '\n')


def scan_awvs(file_list):
    if avsm == True:
        print('开始调用awvs')
        try:
            os.system('rm -rf awvsput.out')
        except:
            print('不存在该文件')
        os.system('nohup python3 awvs_monitor.py >awvsput.out 2>&1 &')
    awvs_file_name = './result/awvslist/all_av_list.txt'
    with open(awvs_file_name, 'a', encoding='utf-8') as f:
        for a in file_list:
            if 'http' in a:
                print(a)
                f.writelines(a + '\n')


def quchong_info_list(all_info_list):
    host_scan_list = []
    new_list = []
    mgwj_list = []
    ld_list = []
    fs_list = []
    for all in all_info_list:
        if all not in new_list:
            new_list.append(all)
    add_list, yt_fofa_info_list, sm_cache_file_list = save_cache(new_list,subfinder_list)
    add_list = list(set(add_list))
    if len(add_list) != 0:
        filename = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + 'all_url_list.txt'
        temp_list = []
        for a in add_list:
            if len(a) < 4:
                continue
            print('1' + a)
            try:
                if '://' not in a:
                    temp_list.append(a)
                else:
                    a1 = a.split('://')[1]
                    print(a1)
                    if ':' in a1:
                        a1 = a1.split(':')[0]
                    if '/' in a1:
                        temp_list.append(a1.split('/')[0])
                    else:
                        temp_list.append(a1)
            except:
                continue
        temp_list2 = list(set(temp_list))
        with open(filename, 'w', encoding='utf-8') as f:
            for s in temp_list2:
                f.writelines(s + '\n')

        file_filter_name = httpx_naabu_scan(filename, sm_cache_file_list)
        if file_filter_name != 1:
            print('ssss')
            print(file_filter_name)
            # 将生成的url放入对应的域名分类文件中,便于进行host碰撞
            process_and_save_urls()
            print('xxxxx')

            if hostm == True:
                host_scan_list = host_collision()

            if ml == True:
                mgwj_list = ml_sm(file_filter_name)

            if fs == True:
                fs_list = fscan(file_filter_name, ip_list)

            if ld == True:
                ld_list = nuclei(file_filter_name)

    # 扫描自己收集的资产
    if notauto == True:
        filename = './result/notautolist/notautolist.txt'
        new_filename = './result/notautolist/' + str(
            time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())) + 'new_notautolist.txt'
        add_list = open(filename, 'r', encoding='utf-8').read().split('\n')
        temp_list = []
        for a in add_list:
            if len(a) < 4:
                continue
            print('1' + a)
            try:
                a1 = a.split('://')[1]
                print(a1)
                if ':' in a1:
                    a1 = a1.split(':')[0]
                if '/' in a1:
                    temp_list.append(a1.split('/')[0])
                else:
                    temp_list.append(a1)

            except:
                if 'http:' not in a and 'htts:' not in a:
                    if ':' in a:
                        a = a.split(':')[0]
                    if '/' in a:
                        temp_list.append(a.split('/')[0])
                    else:
                        temp_list.append(a)
                continue
        temp_list2 = list(set(temp_list))

        with open(new_filename, 'w', encoding='utf-8') as f:
            for s in temp_list2:
                f.writelines(s + '\n')

        file_filter_name = httpx_naabu_scan(new_filename, sm_cache_file_list)
        print('ssss')
        print(file_filter_name)
        print('xxxxx')
        # 添加host碰撞
        if hostm == True:
            host_scan_list = host_collision()

        if ml == True:
            mgwj_list = ml_sm(file_filter_name)

        if ld == True:
            ld_list = nuclei(file_filter_name)

        if fs == True:
            fs_list = fscan(file_filter_name, ip_list)

    print('==============================')
    # print(mgwj_list)
    # print(ld_list)
    print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx')
    print(yt_fofa_info_list)
    print('zzzzzzzzzzzzzzzzzzzzzzzzzzzzzzz')
    return yt_fofa_info_list, mgwj_list, ld_list, fs_list, host_scan_list


def bypass403(mg_url):
    os.system('./inifile/bypass403/f403_linux_amd64 -u ' + mg_url + '> temp.txt')
    l = open('./inifile/bypass403/temp.txt', 'r', encoding='utf-8').read().split('\n')
    l2 = []
    for i in l:
        print(i)
        if '[*]' in i or '[+]' in i:
            l2.append(i)

    pass_url_list = []
    for i in range(0, len(l2)):
        print(l2[i])
        if '[+]' in l2[i]:
            info = []
            url = l2[i].split(' ')[3].split('\x1b')[0] + '==>' + \
                  l2[i - 1].replace(' ', '').split('[*]')[1].split('\x1b')[0]
            code = l2[i].split(' ')[1]
            words = l2[i].split(' ')[2]
            info.append(url)
            info.append(code)
            info.append(words)
            pass_url_list.append(info)

    os.system('rm -rf ./inifile/bypass403/temp.txt')
    return pass_url_list


def host_collision():
    result_list = []
    if len(ip_list) > 1:
        # 添加域名IP到ip_as_domain文件夹
        append_ip_to_domain_file(all_domain_ip_list)
        ip_file_list = process_domian_ip_files('./result/ip_as_domain/')
        print(ip_file_list)
        for ip_file in ip_file_list:
            if 'temp' not in ip_file:
                """ Extracts and returns the collision success data from the provided file. """
                # Read content from the file
                file_path = './inifile/hostscan/' + time.strftime("%Y-%m-%d-%H-%M-%S",
                                                                  time.localtime()) + "_host_scan.txt"
                domain_file = './result/maindomain/' + ip_file.split('./result/ip_as_domain/')[1]
                print(ip_file)
                print(domain_file)
                os.system(
                    "java -jar ./inifile/hostscan/HostCollision.jar -ifp " + ip_file + " -hfp " + domain_file + " -t 10 -cssc 200 > " + file_path)
                print(
                    "java -jar ./inifile/hostscan/HostCollision.jar -ifp " + ip_file + " -hfp " + domain_file + " -t 10 -cssc 200 > " + file_path)

                with open(file_path, 'r', encoding='utf-8') as file:
                    text_content = file.read()

                # Extracting the required portion of the text
                start_keyword = "====================碰 撞 成 功 列 表===================="
                end_keyword = "执行完毕 ヾ(≧▽≦*)o"
                start_index = text_content.find(start_keyword) + len(start_keyword)
                end_index = text_content.find(end_keyword)

                # Extracted text
                extracted_text = text_content[start_index:end_index].strip()

                # Splitting the text into lines and adding to the list
                result = extracted_text.split('\n')
                for res in result:
                    if res != '' and '协议' in res and '匹配失败-3' not in res and 'title:,' not in res:
                        result_list.append(res)

        folder_path = './result/ip_as_domain'
        # 列出文件夹下的所有文件（包括子文件夹）
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            # 如果是文件，则删除
            if os.path.isfile(file_path):
                os.remove(file_path)

        # Example usage
        try:
            for filename in os.listdir('./'):
                file_path = os.path.join('./', filename)
                # Check if it's a file and matches the criteria
                if (os.path.isfile(file_path) and filename.startswith('202') and
                        '_' in filename and (filename.endswith('.txt') or filename.endswith('.csv'))):
                    os.remove(file_path)
        except:
            print('无文件')

        return result_list
    else:
        return result_list


def process_domian_ip_files(directory_path):
    """ List all files in the directory, add them to a list, and then delete them. """
    test = []  # List to hold the file paths

    # Check if the directory exists
    if os.path.exists(directory_path) and os.path.isdir(directory_path):
        # List all files and directories in the specified directory
        for item in os.listdir(directory_path):
            item_path = os.path.join(directory_path, item)
            # Check if it is a file and not a directory
            if os.path.isfile(item_path):
                test.append(item_path)

    else:
        print(f"Directory {directory_path} does not exist.")

    return test


def ml_sm(filename):
    url_list = open(filename, 'r', encoding='utf-8', errors='ignore').read().split('\n')
    # 返回的字节长度列表
    result = []

    for url in url_list:
        msg_info = []
        msg_info2 = []
        print(url)
        try:
            if 'http' in url:
                temp_file = 'temp_result.txt'
                print(
                    './inifile/ffuf/ffuf -u ' + url + '/FUZZ -w ./inifile/dict/file_top_200.txt -ac -t 100 -o ' + temp_file)
                os.system(
                    './inifile/ffuf/ffuf -u ' + url + '/FUZZ -w ./inifile/dict/file_top_200.txt -ac -t 100 -o ' + temp_file)

            else:
                continue

            with open(temp_file, 'r', encoding='utf-8') as f:
                data = json.load(f)['results']
                f.close()

            # 删除临时文件
            os.system('rm -rf ' + temp_file)
            # print(data)

            # 存放返回包长度
            for i in range(len(data)):
                msg_info.append(str(tldextract.extract(data[i]['url']).registered_domain) + str(data[i]['words']))
            for i in range(len(data)):
                msg_info2.append(str(data[i]['words']))
            # print(msg_info)

            for i in range(len(data)):
                info_list = []
                if msg_info.count(
                        str(tldextract.extract(data[i]['url']).registered_domain) + str(data[i]['words'])) == 1 and \
                        data[i]['words'] > 100 and msg_info2.count(str(data[i]['words'])) < 4:
                    info_list.append(data[i]['url'])
                    info_list.append(data[i]['status'])
                    info_list.append(data[i]['words'])
                    result.append(info_list)
                else:
                    print('1')

                # 对扫描的403,401页面进行bypass扫描
                if data[i]['status'] == 403 or data[i]['status'] == 401:
                    l = bypass403(data[i]['url'])
                    if len(l) != 0:
                        for i in l:
                            result.append(i)
        except:
            continue

    return result


def dingtalk(message_list, mgml_list, ld_list, fs_list, host_scan_list):
    # 钉钉WebHook地址
    DWebHook = 'https://oapi.dingtalk.com/robot/send?access_token=' + str(dingding_hook)
    Dsecret = dingding_key  # 可选：创建机器人勾选“加签”选项时使用
    # 初始化机器人小丁
    # xiaoding = DingtalkChatbot(webhook)  # 方式一：通常初始化方式
    xiaoding = DingtalkChatbot(DWebHook, secret=Dsecret)  # 方式二：勾选“加签”选项时使用（v1.5以上新功能）
    # xiaoding = DingtalkChatbot(webhook, pc_slide=True)  # 方式三：设置消息链接在PC端侧边栏打开（v1.5以上新功能）
    # Text消息@所有人
    # xiaoding.send_text(msg=c, is_at_all=True)

    # host碰撞信息
    # fscan扫描漏洞信息
    new_list5 = []
    num5 = len(host_scan_list)
    print(host_scan_list)
    if int(num5) > 100:
        n5 = num5 // 50
    else:
        n5 = 1

    for i in range(n5):
        one_list = host_scan_list[math.floor(i / n5 * num5):math.floor((i + 1) / n5 * num5)]
        new_list5.append(one_list)
    if num5 != 0:
        for i in new_list5:
            xuhao = 1
            message = ''
            title = dingding_tag + ':新收集host碰撞成功信息 ' + str(
                num5) + ' 个' + '\n' + '-----------------------------------------------'

            for msg in i:
                message = message + str(xuhao) + '.' + str(msg) + '\n'
                xuhao += 1
                message = title.lstrip() + '\n' + message
                title = ''
            message + '\n' + '-----------------------------------------------'
            msg5 = message.lstrip('\n')
            if message != '':
                xiaoding.send_text(msg=msg5)

    # fscan扫描漏洞信息
    new_list4 = []
    num4 = len(fs_list)
    print(fs_list)
    if int(num4) > 100:
        n4 = num4 // 50
    else:
        n4 = 1

    for i in range(n4):
        one_list = fs_list[math.floor(i / n4 * num4):math.floor((i + 1) / n4 * num4)]
        new_list4.append(one_list)
    if num4 != 0:
        for i in new_list4:
            xuhao = 1
            message = ''
            title = dingding_tag + ':新收集fscan扫描漏洞信息 ' + str(
                num4) + ' 个' + '\n' + '-----------------------------------------------'

            for msg in i:
                message = message + str(xuhao) + '.' + str(msg) + '\n'
                xuhao += 1
                message = title.lstrip() + '\n' + message
                title = ''
            message + '\n' + '-----------------------------------------------'
            msg4 = message.lstrip('\n')
            if message != '':
                xiaoding.send_text(msg=msg4)

    # 漏洞信息个数
    msg_info = []
    # 漏洞信息
    new_list3 = []
    num3 = len(ld_list)
    print(ld_list)
    if int(num3) > 100:
        n3 = num3 // 50
    else:
        n3 = 1

    for i in range(n3):
        one_list = ld_list[math.floor(i / n3 * num3):math.floor((i + 1) / n3 * num3)]
        new_list3.append(one_list)
    if num3 != 0:
        for i in new_list3:
            xuhao = 1
            message = ''
            title = dingding_tag + ':新收集漏洞信息 ' + str(
                num3) + ' 个,排除一些抽象漏洞后如下' + '\n' + '-----------------------------------------------'

            for msg in i:
                msg_info.append(msg[0])

            for msg in i:
                if 'ssl' in msg[0] or 'tls' in msg[0] or '-certificate' in msg[0] or 'cipher' in msg[
                    0] or 'mismatched' in msg[0] or msg_info.count(msg[0]) > 6:
                    continue
                info = str(msg[0]) + '   ' + str(msg[1]) + '   ' + str(msg[2])
                message = message + str(xuhao) + '.' + str(info) + '\n'
                xuhao += 1
                message = title.lstrip() + '\n' + message
                title = ''
            message + '\n' + '-----------------------------------------------'
            msg3 = message.lstrip('\n')
            if message != '':
                xiaoding.send_text(msg=msg3)

    # 敏感数据
    new_list2 = []
    num2 = len(mgml_list)
    print(mgml_list)
    if int(num2) > 100:
        n2 = num2 // 50
    else:
        n2 = 1

    for i in range(n2):
        one_list = mgml_list[math.floor(i / n2 * num2):math.floor((i + 1) / n2 * num2)]
        new_list2.append(one_list)
    if num2 != 0:
        for i in new_list2:
            xuhao = 1
            message = ''
            title = dingding_tag + ':新收集敏感信息 ' + str(
                num2) + ' 个,其中返回为200的如下' + '\n' + '-----------------------------------------------'

            for msg in i:
                if str(msg[1]) != '200':
                    continue
                info = str(msg[0]) + '   ' + str(msg[1]) + '   ' + str(msg[2])
                message = message + str(xuhao) + '.' + str(info) + '\n'
                xuhao += 1
                message = title.lstrip() + '\n' + message
                title = ''
            message + '\n' + '-----------------------------------------------'
            msg2 = message.lstrip('\n')
            if message != '':
                xiaoding.send_text(msg=msg2)

    # 单独@人员
    new_list = []
    num = len(message_list)
    print(message_list)
    if int(num) > 200:
        n = num // 50
    else:
        n = 1
    for i in range(n):
        one_list = message_list[math.floor(i / n * num):math.floor((i + 1) / n * num)]
        new_list.append(one_list)
    if num != 0:
        for i in new_list:
            print(i)
            xuhao = 1
            num1 = len(i)
            title = dingding_tag + ':新收集暴露面信息 ' + str(
                num1) + ' 个,其中状态码为200的如下' + '\n' + '-----------------------------------------------' + '\n' + '网址           ' + '    状态码    ' + '     标题      '
            message = ''
            for msg in i:
                if str(msg[1]) != '200':
                    continue
                info = str(msg[0]) + '   ' + str(msg[1]) + '   ' + str(msg[2]) + '   '
                message = message + str(xuhao) + '.' + str(info) + '\n'
                xuhao += 1
                message = title.lstrip() + '\n' + message
                title = ''
            message + '\n' + '-----------------------------------------------'
            msg = message.lstrip('\n')
            if message != '':
                xiaoding.send_text(msg=msg)


# 调用搜集的所有IP列表进行fscan扫描
def fscan(filename, ip_list):
    l = open('./result/allip/ip_cache.txt', 'r', encoding='utf-8').read().split('\n')
    url_file = './result/allip/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + '_allip.txt'
    loud_file = './result/fscan/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + '_fscan.txt'
    loud_file2 = './result/fscan/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + '_ip_fscan.txt'
    with open(url_file, 'w') as f:
        for ip in ip_list:
            if ip not in l:
                f.writelines(ip + '\n')
    os.system('./inifile/lousao/fscan  -uf ' + filename + ' -o ' + str(loud_file))
    # os.system('./inifile/lousao/fscan  -p 22,3389,445,3306,1433,1521,21,27017,11211,5432,23,25,465,110,995,143,993,5900,6379 -np -hf ' + url_file + ' -o ' + str(loud_file2))
    os.system('./inifile/lousao/fscan  -np -hf ' + url_file + ' -o ' + str(loud_file2))
    list1 = []
    with open('./result/allip/ip_cache.txt', 'a') as f:
        for ip in ip_list:
            f.writelines(ip + '\n')
    try:
        with open(loud_file, 'r') as f:
            print(loud_file)
            test1 = f.readlines()
            for t in test1:
                if '[+]' in t and '扫描结束' not in t:
                    list1.append(t.strip('\n'))
    except:
        print('无漏洞')
    try:
        with open(loud_file2, 'r') as f:
            print(loud_file2)
            test1 = f.readlines()
            for t in test1:
                if '[+]' in t and '扫描结束' not in t:
                    list1.append(t.strip('\n'))
    except:
        print('无漏洞')

    print(list1)
    return list1


def nuclei(filename):
    os.system('./inifile/lousao/nuclei -update')
    os.system('./inifile/lousao/nuclei -update-templates ')
    loud_file = './result/loudong/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + 'ld_scan.txt'
    # os.system('./inifile/lousao/nuclei -un -ut')
    os.system('./inifile/lousao/nuclei -mhe 3 -timeout 1 -rl 300 -c 50  -s low,medium,high,critical -l ' + str(
        filename) + ' -o ' + str(loud_file))
    list1 = []
    list2 = []
    try:
        with open(loud_file, 'r') as f:
            print(loud_file)
            test1 = f.readlines()
            for t in test1:
                if t[0] != '#' and t[0] != '\n':
                    list1.append(t.strip('\n'))
    except:
        print('该网站无漏洞信息')
        return list2

    for l in list1:
        temp = []
        x = l.split(' ')
        print(x)
        temp.append(x[0])
        temp.append(x[2])
        temp.append(x[3])
        list2.append(temp)
    print(list2)

    return list2


def process_and_save_urls():
    """ Process each URL in the given file, save them into domain-named files, and avoid duplicates. """
    # Ensure the output folder exists
    folder_path = './result/maindomain/'
    os.makedirs(folder_path, exist_ok=True)
    domian_list = []
    for http in httpx_info:
        if '40' in http[1] or '50' in http[1]:
            domian_list.append(http[0])
    print('=-=')
    print(domian_list)
    print('=-=')
    for url in domian_list:
        url = url.strip()
        if url:
            # Remove http:// or https:// from the URL
            url = url.replace('http://', '').replace('https://', '')

            # Extract the main domain using tldextract
            domain = tldextract.extract(url).registered_domain

            if domain:
                # Path for the new file named after the domain
                domain_file_path = os.path.join(folder_path, domain + '.txt')

                # Check if the URL is already in the file
                if not os.path.exists(domain_file_path) or url not in open(domain_file_path).read():
                    # Write the URL to the corresponding domain file
                    with open(domain_file_path, 'a') as domain_file:
                        domain_file.write(url + '\n')


def append_ip_to_domain_file(domain_ip_list):
    folder_path = './result/ip_as_domain/'
    """ Append the IP address to a file named after the domain's top-level domain. """
    # Extract the top-level domain from the given domain
    # Define the filename
    temp_ip_file = './result/ip_as_domain/temp_ip.txt'
    os.system('touch ' + temp_ip_file)
    domain_l = set(domain_ip_list)
    print(domain_l)
    for i in domain_l:
        domain = i.split('-')[0]
        ip = i.split('-')[1]
        tld = tldextract.extract(domain).registered_domain
        filename = os.path.join(folder_path, tld + '.txt')
        os.system('./inifile/naabu/naabu  -host ' + ip + ' -top-ports 1000 -o ' + temp_ip_file)

        # Append or create the file with the IP address
        if os.path.exists(temp_ip_file):
            ip_list = open(temp_ip_file, 'r', encoding='utf-8').read().split('\n')
            for ip in ip_list:
                if not os.path.exists(filename) or ip not in open(filename).read() and len(ip) > 4:
                    with open(filename, 'a') as file:
                        file.write(ip + '\n')


def quchong(l1):
    temp_list = []
    l = []
    for i in l1:
        try:
            temp_list.append(','.join(i))
        except:
            print('jon')
            continue
    print(temp_list)
    l2 = list(set(temp_list))
    print(l2)
    for x in l2:
        l.append(x.split(','))

    return l


def print_custom_help(version, check_msg):
    print(f"""
   __                       
  / /__ ____  __ ____ _____ 
 / / _ `/ _ \/ // / // / -_)
/_/\_,_/\___/\_, /\_,_/\__/ 
            /___/          
    version: {version}({check_msg}) 

项目地址:
https://github.com/Soufaker/laoyue(欢迎issues和star)

项目简介:
一款自动化SRC监控赏金猎人项目，专为安全研究人员和赏金猎人设计，提供多种工具集成和自动化功能，加快漏洞发现和报告流程。

参数说明:
-h, --help  : 显示帮助信息
-d, --domain: 指定单个域名或包含多个域名的文件（如：src.txt）。
-m, --ml    : 使用ffuf扫描目录。可在./inifile/dict目录下添加字典。
-n, --nl    : 使用nuclei进行漏洞扫描。
-f, --fs    : 使用fscan进行漏洞扫描。
-a, --av    : 使用awvs进行漏洞扫描。
-z, --hostz : 进行host碰撞。
-N, --notauto: 启用被动扫描模式，手动收集URL资产后使用(资产放在./result/notautolist/notautolist.txt里)。

常用自动化监控命令(可以先不加nohup手动测试一下看看能跑通不,能跑通就用下面的命令,可以自行增加删除参数,下面的是都跑一遍):
调试命令: python3 laoyue.py -d example.com  -m -f -n -z -a (先跑一个域名,看看能跑通不,能行再使用下面的命令被动监控扫描)
单域名扫描: nohup python3 laoyue.py -d example.com  -m -f -n -z -a  > laoyue.out 2>&1 &
多域名扫描: nohup python3 laoyue.py -d "SRC.txt"  -m -f -n -z -a  > laoyue.out 2>&1 &
被动扫描: nohup python3 laoyue.py -m -n -f -a -N &

额外说明:
1.awvs脚本可单独使用,命令: nohup python3 awvs_monitor.py >awvsput.out 2>&1 &
2.最好运行监控命令: nohup ./check_nohup_size.sh >check_size.out 2>&1 & 
3.目前脚本功能逻辑为: fofa,yt,subfinder(url定期收集)-->naabu(端口扫描)-->httpx(存活探测)-->(对收集的资产的url进行去重,对IP进行cdn检测去重)-->进行host碰撞,漏洞扫描,敏感目录扫描,弱口令探测等-->(归纳信息进行去重)-->本地服务器保存excel-->发送消息(钉钉,企业微信)
4.因为subfinder第一次会跑很多域名,所以第一次启动会比较慢,后续自动化播报就会变快,想要很快的话可以在配置文件去掉subfinder扫描看个人。
5.后续持续添加新功能和优化速度

    """)


# 获取github项目版本
def get_latest_release_version():
    try:
        # GitHub API URL for releases
        api_url = f"https://api.github.com/repos/Soufaker/laoyue/releases/latest"

        # Make a GET request to the GitHub API
        response = requests.get(api_url)
        response.raise_for_status()

        # Parse the JSON response
        data = response.json()

        # Return the tag name (version)
        return data.get("tag_name", "Unknown")
    except Exception as e:
        return "Unknown"


if __name__ == '__main__':
    # 参数设置
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument('-h', '--help', action='store_true', help='显示帮助信息')
    # 添加其他参数
    parser.add_argument('-d', '--domain', action='store', default='', dest="company_domains")
    parser.add_argument('-m', '--ml', action='store_true', dest="ml_sm")
    parser.add_argument('-n', '--nl', action='store_true', dest="ld_sm")
    parser.add_argument('-f', '--fs', action='store_true', dest="fs_sm")
    parser.add_argument('-a', '--av', action='store_true', dest="av_sm")
    parser.add_argument('-z', '--hostz', action='store_true', dest="host_sm")
    parser.add_argument('-N', '--notauto', action='store_true', dest="not_auto")

    args = parser.parse_args()

    # 加载配置文件
    cf = ConfigParser()
    cf.read('./config/config.ini')
    global hunter_config_list
    hunter_config_list = []
    global fofa_key
    fofa_key = ''
    global fofa_email
    fofa_email = ''
    global black_domian
    black_domian = []
    global dingding_key
    dingding_key = ''
    global dingding_hook
    dingding_hook = ''
    global fofa_size
    fofa_size = ''
    global fofa_keyword
    fofa_keyword = ''
    global yt_size
    yt_size = ''
    global yt_keword
    yt_keword = ''
    global fofa_count
    fofa_count = ''
    global hunter_count
    hunter_count = ''
    global is_subfinder
    is_subfinder = ''
    global is_fofa
    is_fofa = ''
    global is_hunter
    is_hunter = ''
    global file_filter_name
    file_filter_name = ''
    global httpx_info
    httpx_info = []
    global dingding_tag
    dingding_tag = ''

    c_len = cf.options('hunter')
    for i in c_len:
        hunter_config_list.append(cf.get('hunter', i))
    yt_keword = ''
    fofa_count = cf.get('fofa', 'count')
    hunter_count = cf.get('hunter', 'count')
    fofa_size = cf.get('fofa', 'size')
    is_fofa = cf.get('fofa', 'is_fofa')
    is_subfinder = cf.get('subfinder', 'is_subfinder')
    is_hunter = cf.get('hunter', 'is_hunter')
    yt_size = cf.get('hunter', 'size')
    fofa_keyword = cf.get('fofa', 'keyword')
    yt_keword = cf.get('hunter', 'keyword')
    fofa_key = cf.get('fofa', 'fofa_key')
    fofa_email = cf.get('fofa', 'fofa_email')
    dingding_hook = cf.get('dingding', 'access_token')
    dingding_key = cf.get('dingding', 'dsecret')
    black = cf.get('black_domain', 'domain')
    black_domian = black.split(',')
    dingding_tag = cf.get('tag', 'dingding_tag')
    # 所有搜集到的URL列表
    all_url_list = []

    # 最后整理出的url列表
    all_url_list2 = []

    # 全局收集的信息
    global all_info_list
    all_info_list = []

    # 全局搜寻的域名
    all_domain_list = []

    # 全局域名和IP的映射表
    global all_domain_ip_list
    all_domain_ip_list = []

    # subfinder扫描的列表
    global subfinder_list
    subfinder_list = []

    # 初始化
    quchong_list = []
    mgwj_list = []
    ld_list = []
    fs_list = []
    global ip_list
    ip_list = []

    # 参数获取
    global ml
    ml = args.ml_sm

    global ld
    ld = args.ld_sm

    global fs
    fs = args.fs_sm

    global avsm
    avsm = args.av_sm

    global hostm
    hostm = args.host_sm

    global notauto
    notauto = args.not_auto

    global company_domain
    company_domain = args.company_domains

    # 调用fofa,yt获取信息
    if args.help:
        check_msg = ''
        now_version = 'v1.2.3'
        new_version = get_latest_release_version()
        if now_version == new_version:
            check_msg = '现在是最新版本'
        else:
            check_msg = '新版本:' + new_version + ',请自行进行更新'

        print_custom_help(now_version, check_msg)
    else:
        # 在这里处理其他参数
        get_all_url_fo_yt()
        quchong_list, mgwj_list, ld_list, fs_list, host_scan_list = quchong_info_list(all_info_list)

        if len(quchong_list) != 0 or len(mgwj_list) != 0 or len(ld_list) != 0 or len(
                fs_list) != 0 or len(host_scan_list) != 0:
            Write_To_Excel(quchong_list, mgwj_list, ld_list, httpx_info, fs_list, host_scan_list)
            # 发送信息
            try:
                set_info = quchong(httpx_info)
                dingtalk(set_info, mgwj_list, ld_list, fs_list, host_scan_list)
            except:
                print('发送消息异常')
                traceback.print_exc()
                time.sleep(60)
                try:
                    dingtalk(set_info, mgwj_list, ld_list, fs_list, host_scan_list)
                except:
                    print('网络存在问题,继续执行任务')

        if notauto != True:
            time.sleep(360)
            try:
                os.system('rm -rf laoyue.out')
                os.system('nohup python3 laoyue.py -d "SRC.txt"  -m -f -n -z -a  > laoyue.out 2>&1 &')
            except:
                print('laoyue.out文件不存在')

