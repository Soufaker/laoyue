#author:soufaker
#time:2021/12/14

import requests
import optparse
from urllib.parse import quote,unquote
from lxml import etree
from openpyxl import Workbook
import time

# 选取想要的域名元素
def Get_MiddleStr(content,startStr,endStr,num): #获取中间字符串的⼀个通⽤函数
    try:
        startIndex = content.index(startStr)
        if startIndex>=0:
            if num != 1:
                endIndex = startIndex
                startIndex = endIndex + num

            else:
                startIndex += len(startStr)
                endIndex = content.index(endStr)

        return content[startIndex:endIndex]

    except:
        return 'ok'

# 获取公司查询ID
def Get_Company_Id(company_name,company_name_list):
    company_id_list = []
    url = "https://www.tianyancha.com/search?key="
    endstr = "target='_blank'>"
    starstr = "https://www.tianyancha.com/company/"
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    if company_name != None:
        response = requests.get(url=url + quote(company_name, 'utf-8'), headers=header).text
        resp = response.replace(' ', '').replace('\n', '').replace('\t', '').replace("\"", "")
        company_id_list.append(Get_MiddleStr(resp, starstr, endstr,1))
    else:
        with open(company_name_list,'r',encoding='utf-8') as f:
            for company in f:
                response = requests.get(url=url+quote(company,'utf-8'),headers=header).text
                resp = response.replace(' ', '').replace('\n', '').replace('\t', '').replace("\"", "")
                company_id_list.append(Get_MiddleStr(resp,starstr,endstr,1))

    return company_id_list

#获取网页股权架构图页面数量
def Get_Page_Num(company_id):

    company_url = "https://www.tianyancha.com/company/" + company_id
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    response = requests.get(url=company_url, headers=header).text
    resp = response.replace(' ', '').replace('\n', '').replace('\t', '').replace("\"", "")
    page_num = Get_MiddleStr(resp,'num-endonclick=companyPageChange(',',this',3)
    if page_num == 'ok':
        return 'ok'
    else:
        if page_num[len(page_num) - 1:] == ',':
            page_real_num = page_num[0:2]
        else:
            page_real_num = page_num[0:1]

        return page_real_num

#获取公司占百分百股权且未注销的子公司域名信息
def Get_ALL_Sub_Cmpany_Domain(company_id):
    company_info_list = []
    company_url_list = []
    for main_company in company_id:
        l,l2 = Get_Sub_Company_Domain(main_company)
        if len(l) != 0:
            for i in l:
                company_info_list.append(i)
        if len(l2) != 0:
            for i2 in l2:
                company_url_list.append(i2)

        com_id = []
        com_id.append(main_company)
        while com_id:
            co_id = com_id.pop(0)
            company_id_list = get_all_page_id(co_id)
            if len(company_id_list) != 0:
                for co in company_id_list:
                    com_id.append(co)
                for ci in  company_id_list:
                    l, l2 = Get_Sub_Company_Domain(ci)
                    if len(l) != 0:
                        for i in l:
                            company_info_list.append(i)
                    if len(l2) != 0:
                        for i2 in l2:
                            company_url_list.append(i2)

    return company_info_list,company_url_list

def Get_Sub_Company_Domain(company_id):
    company_info_list = []
    company_url_list = []
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    company_url = "https://www.tianyancha.com/company/" + company_id
    print(company_url)
    try:
        res = requests.get(url=company_url, headers=header)
        response2 = res.content
        html = etree.HTML(response2)
        name = html.xpath('/html/body/div[2]/div[2]/div[2]/div/div/div[3]/div[3]/div[1]/span/span/h1/text()')
        url = html.xpath('/html/body/div[2]/div[2]/div[2]/div/div/div[3]/div[3]/div[3]/div[3]/div[1]/a[2]/text()')
        email = html.xpath(
            '/html/body/div[2]/div[2]/div[2]/div/div/div[3]/div[3]/div[3]/div[2]/div[2]/span/span[4]/text()')

        # 对存在未空的url切换标签
        if len(url) == 0:
            url.append('暂无网址')
        elif len(email) == 0:
            email.append('暂无邮箱')
        else:
            company_url_list.append(url[0])

        info = [str(name[0]), str(url[0]), str(email[0])]
        company_info_list.append(info)


    except:
        time.sleep(5)
        print('页面访问过快,无法获取数据!,请切换代理IP')

    return company_info_list,company_url_list

def Get_Page_Num(company_id):
    company_url = "https://www.tianyancha.com/company/" + str(company_id)
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    response = requests.get(url=company_url, headers=header).text
    resp = response.replace(' ', '').replace('\n', '').replace('\t', '').replace("\"", "")
    page_num = Get_MiddleStr(resp,'</a></li><li><aclass=num-nextonclick=companyPageChange(','',-2)
    if page_num[0:1] == '>':
        return page_num[1:]
    else:
        return page_num

def get_all_page_id(num):
    company_oc = company_occ
    occ_list = []
    company_url_list = []
    id_list = []
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    with open('cookie.ini', 'r', encoding='utf-8') as co:
        f = co.read()
        f2 = f.replace(' ','')
        header['Cookie'] = f2[6:]
    page_num = Get_Page_Num(num)
    if page_num == 'ok':
        company_url_list.append("https://www.tianyancha.com/pagination/investV2.xhtml?&pn=1&id="+str(num))
    else:
        for p in range(1,int(page_num)):
            company_url_list.append("https://www.tianyancha.com/pagination/investV2.xhtml?ps=20&pn="+str(p)+"&id="+str(num))
        company_url_list.append("https://www.tianyancha.com/pagination/investV2.xhtml?&pn="+str(page_num)+"&id="+str(num))
    print(company_url_list)
    for c in company_url_list:
        response = requests.get(url=c, headers=header).text
        resp = response.replace(' ', '').replace('\n', '').replace('\t', '').replace("\"", "")
        resp_index = resp.index('序号')
        resp_filter = resp[resp_index:]
        resp_list = resp_filter.split('%')
        p_index = 0
        for res in resp_list:
            p_index = p_index + int(len(res)) + 1
            res_re = res + '%'
            res_rev = res_re[::-1]
            occ = Get_MiddleStr(res_rev, '%', '>', 1)
            occ_r = occ[::-1]
            occ_list.append(occ_r)
            if occ_r != '':
                occ_rev = float(occ_r)
                if occ_rev >= company_oc and resp_filter[p_index + 24:p_index + 29] != 'cance':
                    id = Get_MiddleStr(res, 'https://www.tianyancha.com/company/', 'target=_blank>', 1)
                    if id != 'ok':
                        id_list.append(id)
    return id_list

def Write_To_Excel(company_info_list):
    t = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
    wb = Workbook()
    ws = wb.active

    ws['A1'] = '公司名'
    ws['B1'] = '网址'
    ws['C1'] = '邮箱'
    for l in company_info_list:
        ws.append(list(l))

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 36
    wb.save("company_info_list"+t+".xlsx")


def Write_To_Txt(company_url_list):
    t = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
    with open('company_url_list'+t+'.txt','w+',encoding='utf-8') as f:
        for c in company_url_list:
            f.writelines(c+'\n')

if __name__ == '__main__':
    # 参数设置
    parser = optparse.OptionParser()
    parser.add_option('-c', '--company', action='store', dest="company_name")
    parser.add_option('-l', '--list', action='store', dest="company_name_list")
    parser.add_option('-o', '--occ', action='store',default='100', dest="company_occ")
    options, args = parser.parse_args()

    # 参数获取
    global company_occ
    company_name = options.company_name
    company_list = options.company_name_list
    company_occ = float(options.company_occ)
    company_id = Get_Company_Id(company_name, company_list)
    company_info_list, company_url_list = Get_ALL_Sub_Cmpany_Domain(company_id)
    Write_To_Txt(company_url_list)
    Write_To_Excel(company_info_list)
